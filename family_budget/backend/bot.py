import os
import logging
import asyncio
import threading
import uuid
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, ContextTypes, MessageHandler, CommandHandler, CallbackQueryHandler, filters
from sqlmodel import Session
from backend.database import engine, Expense, ActivityLog
from backend.parser import parse_expense_text, parse_bulk_transactions

logger = logging.getLogger(__name__)


def log_activity(
    action: str,
    source: str,
    details: str = "",
    record_count: int = None,
    file_date: str = None,
    total_amount: float = None
):
    """Log an activity to the activity_logs table."""
    import json
    from datetime import date as date_type

    with Session(engine) as session:
        log_entry = ActivityLog(
            action=action,
            source=source,
            details=details if isinstance(details, str) else json.dumps(details, ensure_ascii=False),
            record_count=record_count,
            file_date=date_type.fromisoformat(file_date) if file_date else None,
            total_amount=total_amount
        )
        session.add(log_entry)
        session.commit()


# In-memory storage for pending bulk imports (keyed by unique ID)
# Format: {import_id: {"transactions": [...], "sender_name": str, "chat_id": int}}
pending_imports = {}

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
ALLOWED_USER_IDS = [
    int(uid.strip())
    for uid in os.getenv("ALLOWED_USER_IDS", "").split(",")
    if uid.strip().isdigit()
]


def is_authorized(user_id: int, chat_id: int) -> bool:
    """Check if user/chat is authorized to use the bot."""
    if not ALLOWED_USER_IDS:
        # No whitelist configured = allow all
        return True
    return user_id in ALLOWED_USER_IDS or chat_id in ALLOWED_USER_IDS


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command."""
    if not update.message:
        return

    user = update.effective_user
    chat_id = update.effective_chat.id

    if not is_authorized(user.id, chat_id):
        logger.warning(f"Unauthorized /start from user ID: {user.id}")
        return

    welcome = (
        "Family Budget Bot\n\n"
        "Send me expenses in natural language:\n"
        "• `coffee 25`\n"
        "• `groceries 340`\n"
        "• `150 fuel`\n"
        "• `wolt pizza 89`\n\n"
        "I'll auto-classify and log them!"
    )
    await update.message.reply_text(welcome, parse_mode="Markdown")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /help command."""
    if not update.message:
        return

    user = update.effective_user
    chat_id = update.effective_chat.id

    if not is_authorized(user.id, chat_id):
        return

    help_text = (
        "*Quick Expense Logging*\n\n"
        "Just type the expense naturally:\n"
        "• `coffee 18` → Restaurants & Dining\n"
        "• `shufersal 250` → Groceries & Supermarket\n"
        "• `parking 30` → Transportation & Fuel\n\n"
        "*Bulk Import:*\n"
        "• `/bulk` followed by multiple lines\n"
        "• Or upload a file (PDF, TXT, CSV, XLSX)\n\n"
        "*Supported Keywords:*\n"
        "Groceries: supermarket, groceries, shufersal, rami levy, mega\n"
        "Transport: fuel, gas, charging, parking, sonol, paz\n"
        "Dining: coffee, restaurant, wolt, cafe, pizza\n\n"
        "Unrecognized keywords go to 'General & Miscellaneous'.\n\n"
        "View dashboard at: http://localhost:8000"
    )
    await update.message.reply_text(help_text, parse_mode="Markdown")


async def bulk_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /bulk command - parse multi-line transactions."""
    if not update.message:
        return

    user = update.effective_user
    chat_id = update.effective_chat.id

    if not is_authorized(user.id, chat_id):
        return

    # Get text after the /bulk command
    raw_text = update.message.text
    if " " in raw_text:
        bulk_text = raw_text.split(" ", 1)[1]
    else:
        await update.message.reply_text(
            "*Bulk Import*\n\n"
            "Paste your CC transactions after the command:\n"
            "`/bulk\n"
            "Shufersal 150.00\n"
            "Wolt 89.50\n"
            "Paz Gas 220.00`\n\n"
            "Or upload a file (PDF, TXT, CSV, XLSX) with your statement.",
            parse_mode="Markdown"
        )
        return

    await process_bulk_text(update, bulk_text, user.first_name or "Family Member")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads (PDF, TXT, CSV) for bulk transaction import."""
    if not update.message or not update.message.document:
        return

    user = update.effective_user
    chat_id = update.effective_chat.id

    if not is_authorized(user.id, chat_id):
        logger.warning(f"Unauthorized document from user ID: {user.id}")
        return

    document = update.message.document
    file_name = document.file_name or "unknown"
    file_size = document.file_size or 0

    # Limit file size (1MB max)
    if file_size > 1024 * 1024:
        await update.message.reply_text("File too large. Maximum size is 1MB.")
        return

    # Check file type
    allowed_extensions = [".txt", ".csv", ".pdf", ".xlsx", ".xls"]
    file_ext = "." + file_name.split(".")[-1].lower() if "." in file_name else ""

    if file_ext not in allowed_extensions:
        await update.message.reply_text(
            f"Unsupported file type: {file_ext}\n"
            f"Supported formats: TXT, CSV, PDF, XLSX"
        )
        return

    await update.message.reply_text(f"Processing {file_name}...")

    try:
        # Download file
        file = await document.get_file()
        file_bytes = await file.download_as_bytearray()

        # Extract text based on file type
        if file_ext == ".pdf":
            text = extract_text_from_pdf(file_bytes)
        elif file_ext in [".xlsx", ".xls"]:
            text = extract_text_from_xlsx(file_bytes)
        else:
            # TXT or CSV - decode as text
            text = file_bytes.decode("utf-8", errors="ignore")

        if not text.strip():
            await update.message.reply_text("Could not extract text from the file.")
            return

        # Extract expected total for validation
        expected_total = 0.0
        charge_date = ""
        if file_ext == ".pdf":
            expected_total = extract_expected_total_from_pdf(bytes(file_bytes))
            charge_date = extract_charge_date_from_pdf(bytes(file_bytes))
        elif file_ext in [".xlsx", ".xls"]:
            expected_total = extract_expected_total_from_xlsx(bytes(file_bytes))
            charge_date = extract_charge_date_from_xlsx(bytes(file_bytes))

        logger.info(f"Expected total from file: {expected_total}, Charge date: {charge_date}")
        await process_bulk_text(update, text, user.first_name or "Family Member", expected_total, charge_date)

    except Exception as e:
        logger.error(f"Document processing error: {e}")
        await update.message.reply_text(f"Error processing file: {str(e)}")


def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extract text from PDF bytes. Falls back to empty if PyMuPDF not installed."""
    try:
        import pymupdf
        doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except ImportError:
        # Try legacy import for older versions
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except ImportError:
            logger.warning("PyMuPDF not installed - PDF parsing unavailable")
            return ""
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return ""


def extract_expected_total_from_xlsx(xlsx_bytes: bytes) -> float:
    """
    Extract the expected total from an XLSX CC statement.
    Looks for patterns like:
    - MAX: Footer rows with "סך הכל" followed by amount
    - CAL: Header with "עסקאות לחיוב ב-XX/XX/XXXX: X,XXX.XX ₪"
    Returns 0.0 if no total found.
    """
    import re
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        workbook = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
        totals = []

        for sheet in workbook.worksheets:
            all_rows = list(sheet.iter_rows(values_only=True))

            # Check header rows (CAL format: "עסקאות לחיוב ב-XX/XX/XXXX: X,XXX.XX ₪")
            for row in all_rows[:5]:
                row_str = " ".join([str(c) for c in row if c])
                match = re.search(r'עסקאות\s*לחיוב.*?:\s*([\d,]+\.?\d*)\s*₪', row_str)
                if match:
                    amount_str = match.group(1).replace(",", "")
                    totals.append(float(amount_str))

            # Check footer rows (MAX format: "סך הכל" then amount on next row)
            for i, row in enumerate(all_rows):
                row_str = " ".join([str(c) for c in row if c])
                if "סך הכל" in row_str:
                    # Check same row for amount
                    match = re.search(r'([\d,]+\.?\d*)\s*₪', row_str)
                    if match:
                        amount_str = match.group(1).replace(",", "")
                        totals.append(float(amount_str))
                    # Check next row
                    elif i + 1 < len(all_rows):
                        next_row_str = " ".join([str(c) for c in all_rows[i + 1] if c])
                        match = re.search(r'([\d,]+\.?\d*)\s*₪?', next_row_str)
                        if match:
                            amount_str = match.group(1).replace(",", "")
                            try:
                                totals.append(float(amount_str))
                            except ValueError:
                                pass

        workbook.close()
        return sum(totals) if totals else 0.0
    except Exception as e:
        logger.error(f"Error extracting expected total from XLSX: {e}")
        return 0.0


def extract_expected_total_from_pdf(pdf_bytes: bytes) -> float:
    """
    Extract the expected total from a PDF CC statement.
    Looks for patterns like "סה\"כ חיוב לתאריך XX/XX/XX XXXX.XX"
    Returns 0.0 if no total found.
    """
    import re
    try:
        import pymupdf
        doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()

        totals = []

        # Pattern 1: "סה"כ חיוב לתאריך DD/MM/YY AMOUNT" - amount comes AFTER the date
        # The date pattern is DD/MM/YY or DD/MM/YYYY, followed by whitespace/newlines, then the amount
        matches = re.findall(
            r'סה[״"\']כ[\s\n]*חיוב[\s\n]*לתאריך[\s\n]*\d{1,2}/\d{1,2}/\d{2,4}[\s\n]+(\d[\d,]*\.?\d*)',
            full_text, re.DOTALL
        )
        for m in matches:
            try:
                amount = float(m.replace(",", ""))
                if amount > 10:  # Filter out small numbers
                    totals.append(amount)
            except ValueError:
                pass

        # Pattern 2: Amount before "סה"כ" (some formats)
        matches2 = re.findall(r'(\d[\d,]*\.?\d*)\s*₪?\s*סה[״"\']כ', full_text)
        for m in matches2:
            try:
                amount = float(m.replace(",", ""))
                if amount > 10:
                    totals.append(amount)
            except ValueError:
                pass

        # Pattern 3: Generic "סך הכל" followed by amount (with newlines)
        matches3 = re.findall(r'סך[\s\n]*הכל[\s\n]+(\d[\d,]*\.?\d*)', full_text)
        for m in matches3:
            try:
                amount = float(m.replace(",", ""))
                if amount > 10:
                    totals.append(amount)
            except ValueError:
                pass

        return sum(totals) if totals else 0.0
    except ImportError:
        return 0.0
    except Exception as e:
        logger.error(f"Error extracting expected total from PDF: {e}")
        return 0.0


def extract_charge_date_from_pdf(pdf_bytes: bytes) -> str:
    """
    Extract the charge date from PDF CC statement.
    Looks for patterns like "עסקאות לחיוב ב-DD/MM/YYYY" or "חיוב לתאריך DD/MM/YYYY".
    Returns date in YYYY-MM-DD format or empty string if not found.
    """
    import re
    try:
        import pymupdf
        doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()

        # Pattern 1: "עסקאות לחיוב ב-DD/MM/YYYY"
        match = re.search(r'עסקאות\s*לחיוב\s*ב[־-]?\s*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', full_text)
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"

        # Pattern 2: "חיוב לתאריך DD/MM/YYYY"
        match = re.search(r'חיוב\s*לתאריך\s*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', full_text)
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"

        # Pattern 3: "לחיוב בתאריך DD/MM/YYYY"
        match = re.search(r'לחיוב\s*בתאריך\s*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', full_text)
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"

        return ""
    except ImportError:
        return ""
    except Exception as e:
        logger.error(f"Error extracting charge date from PDF: {e}")
        return ""


def extract_charge_date_from_xlsx(xlsx_bytes: bytes) -> str:
    """
    Extract the charge date from XLSX CC statement header.
    Supports multiple formats:
    1. "עסקאות לחיוב ב-DD/MM/YYYY" pattern (Isracard/CAL)
    2. "MM/YYYY" period format (MAX) - returns first day of month
    3. Per-row "תאריך חיוב" column (returns first non-empty value)
    Returns date in YYYY-MM-DD format or empty string if not found.
    """
    import re
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        workbook = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)

        for sheet in workbook.worksheets:
            charge_date_col_idx = None

            # Check first 10 rows for the charge date header
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
                row_str = " ".join([str(c) if c else "" for c in row])

                # Pattern 1: "עסקאות לחיוב ב-DD/MM/YYYY" or "עסקאות לחיוב ב-DD-MM-YYYY"
                match = re.search(r'עסקאות\s*לחיוב\s*ב[־-]?\s*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', row_str)
                if match:
                    day, month, year = match.groups()
                    if len(year) == 2:
                        year = "20" + year
                    charge_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    logger.info(f"Extracted charge date from XLSX (pattern 1): {charge_date}")
                    workbook.close()
                    return charge_date

                # Pattern 2: "MM/YYYY" period format (MAX card exports) - standalone cell
                for cell in row:
                    if cell:
                        cell_str = str(cell).strip()
                        period_match = re.match(r'^(\d{1,2})/(\d{4})$', cell_str)
                        if period_match:
                            month, year = period_match.groups()
                            # Use first day of month as charge date
                            charge_date = f"{year}-{month.zfill(2)}-01"
                            logger.info(f"Extracted charge date from XLSX (period format): {charge_date}")
                            workbook.close()
                            return charge_date

                # Check if this row is a header row with "תאריך חיוב" column
                for col_idx, cell in enumerate(row):
                    if cell and "תאריך חיוב" in str(cell):
                        charge_date_col_idx = col_idx
                        break

            # Pattern 3: If we found a תאריך חיוב column, get the first data value
            if charge_date_col_idx is not None:
                for row in sheet.iter_rows(min_row=5, max_row=20, values_only=True):
                    if len(row) > charge_date_col_idx and row[charge_date_col_idx]:
                        date_val = str(row[charge_date_col_idx])
                        # Parse DD-MM-YYYY or DD/MM/YYYY
                        match = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', date_val)
                        if match:
                            day, month, year = match.groups()
                            if len(year) == 2:
                                year = "20" + year
                            charge_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            logger.info(f"Extracted charge date from XLSX (column): {charge_date}")
                            workbook.close()
                            return charge_date

        workbook.close()
        return ""
    except Exception as e:
        logger.error(f"Error extracting charge date from XLSX: {e}")
        return ""


def extract_text_from_xlsx(xlsx_bytes: bytes) -> str:
    """
    Extract text from Excel (XLSX) bytes.
    Smart handling for Israeli credit card statements which have multiple amount columns.
    Specifically extracts only the charge amount (סכום חיוב) not the original amount (סכום עסקה מקורי).
    Also includes the header row with charge date for GPT to parse.
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        workbook = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
        lines = []

        for sheet in workbook.worksheets:
            # First, capture header rows that may contain charge date info (but not transaction data)
            header_lines = []
            for row_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), start=1):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                row_text = " ".join(row_values).strip()
                # Look for charge date headers like "עסקאות לחיוב ב-" but NOT transaction rows
                # Transaction rows have dates like DD-MM-YYYY at the start
                if row_text and ("עסקאות לחיוב" in row_text or "לתאריך" in row_text):
                    header_lines.append(row_text)

            # Add header lines to output first
            lines.extend(header_lines)

            # Now detect the data columns
            header_row = None
            charge_amount_col = None
            original_amount_col = None

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                row_text = " ".join(row_values)

                # Look for Israeli CC header row with "סכום חיוב"
                if "סכום חיוב" in row_text or "סכום העסקה" in row_text:
                    header_row = row_idx
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell) if cell else ""
                        if "סכום חיוב" in cell_str:
                            charge_amount_col = col_idx
                        elif "סכום עסקה מקורי" in cell_str or "סכום מקורי" in cell_str:
                            original_amount_col = col_idx
                    break

            # If we detected an Israeli CC format, extract smartly
            if header_row and charge_amount_col is not None:
                logger.info(f"Detected Israeli CC format. Charge col: {charge_amount_col}, Original col: {original_amount_col}")

                for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    if not any(cell is not None for cell in row):
                        continue

                    # Build cleaned row: include all columns EXCEPT the original amount column
                    cleaned_values = []
                    for col_idx, cell in enumerate(row):
                        # Skip the original amount column to avoid confusion
                        if original_amount_col is not None and col_idx == original_amount_col:
                            continue
                        # Also skip the currency column right after original amount
                        if original_amount_col is not None and col_idx == original_amount_col + 1:
                            if cell and "₪" in str(cell):
                                continue

                        if cell is not None:
                            cleaned_values.append(str(cell))

                    if cleaned_values:
                        lines.append(" ".join(cleaned_values))
            else:
                # Fallback to generic extraction
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(v.strip() for v in row_values):
                        lines.append(" ".join(row_values))

        workbook.close()

        # Post-process: Remove summary/total lines that confuse GPT into stopping early
        # Keep transaction lines, remove "סך הכל" and standalone amount lines
        import re
        cleaned_lines = []
        for line in lines:
            line_stripped = line.strip()
            # Skip empty lines
            if not line_stripped:
                continue
            # Skip "סך הכל" summary lines
            if line_stripped == "סך הכל":
                continue
            # Skip standalone amount lines like "7937.21₪" or "2681.81₪"
            if re.match(r'^[\d,]+\.?\d*\s*₪?$', line_stripped):
                continue
            cleaned_lines.append(line)

        return "\n".join(cleaned_lines)
    except ImportError:
        logger.warning("openpyxl not installed - XLSX parsing unavailable")
        return ""
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
        return ""


def save_transactions_to_db(transactions: list, sender_name: str, charge_date: str = "") -> int:
    """Save transactions to database and return count.

    Args:
        transactions: List of parsed transactions
        sender_name: Name of the user who uploaded the file
        charge_date: Statement charge date in YYYY-MM-DD format (applies to all transactions)
    """
    from datetime import date as date_type
    saved_count = 0

    # Parse the statement-level charge date
    parsed_charge_date = None
    if charge_date:
        try:
            parsed_charge_date = date_type.fromisoformat(charge_date)
        except (ValueError, TypeError):
            pass

    with Session(engine) as session:
        for tx in transactions:
            # Parse transaction_date if provided (original purchase date)
            tx_date = None
            if tx.get("transaction_date"):
                try:
                    tx_date = date_type.fromisoformat(tx["transaction_date"])
                except (ValueError, TypeError):
                    pass  # Keep as None if parsing fails

            # Parse installment info
            installment_number = tx.get("installment_current")
            total_installments = tx.get("installment_total")
            original_amount = tx.get("original_amount")

            expense = Expense(
                amount=tx["amount"],
                description=tx["description"],
                category_id=tx["category_id"],
                payer=sender_name,
                is_fixed=False,
                source="file",
                transaction_date=tx_date,
                charge_date=parsed_charge_date,
                original_amount=original_amount,
                installment_number=installment_number,
                total_installments=total_installments
            )
            session.add(expense)
            saved_count += 1
        session.commit()
    return saved_count


def build_import_summary(transactions: list, expected_total: float = 0.0) -> tuple[str, str, float]:
    """Build summary text for bulk import. Returns (category_summary, validation_msg, diff_pct)."""
    total = sum(tx["amount"] for tx in transactions)
    categories = {}
    for tx in transactions:
        cat = tx["category_name"]
        categories[cat] = categories.get(cat, 0) + tx["amount"]

    category_summary = "\n".join([f"• {cat}: ₪{amt:,.2f}" for cat, amt in categories.items()])

    validation_msg = ""
    diff_pct = 0.0
    if expected_total > 0:
        diff = abs(total - expected_total)
        diff_pct = (diff / expected_total) * 100 if expected_total > 0 else 0

        if diff_pct <= 1:
            validation_msg = f"\n\n✅ *Validation: MATCH*\nFile total: ₪{expected_total:,.2f}"
        elif diff_pct <= 5:
            validation_msg = f"\n\n⚠️ *Validation: Close*\nFile total: ₪{expected_total:,.2f}\nDifference: ₪{diff:,.2f} ({diff_pct:.1f}%)"
        else:
            validation_msg = f"\n\n❌ *Validation: MISMATCH*\nFile total: ₪{expected_total:,.2f}\nParsed total: ₪{total:,.2f}\nDifference: ₪{diff:,.2f} ({diff_pct:.1f}%)"

    return category_summary, validation_msg, diff_pct


async def process_bulk_text(update: Update, text: str, sender_name: str, expected_total: float = 0.0, charge_date: str = ""):
    """Process bulk transaction text. On mismatch, ask for user confirmation before saving."""
    # Check if OpenAI is configured
    if not OPENAI_API_KEY:
        await update.message.reply_text(
            "⚠️ OpenAI API key not configured.\n"
            "Bulk import requires OpenAI for parsing.\n"
            "Please add openai_api_key in add-on configuration."
        )
        return

    transactions, error = parse_bulk_transactions(text)

    if not transactions:
        # Provide more helpful debugging info
        text_preview = text[:200].replace('\n', ' ')
        error_msg = f"\n\nError: {error}" if error else ""
        await update.message.reply_text(
            f"No transactions found in the text.{error_msg}\n\n"
            f"Text preview: {text_preview}...\n\n"
            f"Make sure the file contains transaction data with amounts."
        )
        return

    total = sum(tx["amount"] for tx in transactions)
    category_summary, validation_msg, diff_pct = build_import_summary(transactions, expected_total)

    # Add charge date info to summary if available
    charge_date_msg = ""
    if charge_date:
        charge_date_msg = f"\nCharge Date: {charge_date}\n"

    # If mismatch (>5%), ask for confirmation before saving
    if expected_total > 0 and diff_pct > 5:
        # Store pending import
        import_id = str(uuid.uuid4())[:8]
        pending_imports[import_id] = {
            "transactions": transactions,
            "sender_name": sender_name,
            "chat_id": update.effective_chat.id,
            "expected_total": expected_total,
            "charge_date": charge_date
        }

        # Build preview message with inline keyboard
        preview_reply = (
            f"*Bulk Import Preview*\n\n"
            f"Transactions: {len(transactions)}\n"
            f"Parsed Total: ₪{total:,.2f}{charge_date_msg}\n\n"
            f"*By Category:*\n{category_summary}"
            f"{validation_msg}\n\n"
            f"⚠️ *There is a significant mismatch!*\n"
            f"Do you want to save these transactions anyway?"
        )

        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ אשר ושמור", callback_data=f"import_approve_{import_id}"),
                InlineKeyboardButton("❌ בטל", callback_data=f"import_decline_{import_id}")
            ]
        ])

        await update.message.reply_text(preview_reply, parse_mode="Markdown", reply_markup=keyboard)
        return

    # No mismatch or mismatch <= 5% - save directly
    saved_count = save_transactions_to_db(transactions, sender_name, charge_date)

    # Log the activity
    log_activity(
        action="file_import",
        source="telegram",
        details=f"Imported by {sender_name}",
        record_count=saved_count,
        file_date=charge_date if charge_date else None,
        total_amount=total
    )

    reply = (
        f"*Bulk Import Complete*\n\n"
        f"Transactions: {saved_count}\n"
        f"Total: ₪{total:,.2f}{charge_date_msg}\n\n"
        f"*By Category:*\n{category_summary}"
        f"{validation_msg}"
    )
    await update.message.reply_text(reply, parse_mode="Markdown")


async def handle_import_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle approve/decline callbacks for bulk imports."""
    query = update.callback_query
    await query.answer()

    data = query.data
    if not data.startswith("import_"):
        return

    parts = data.split("_")
    if len(parts) != 3:
        return

    action = parts[1]  # "approve" or "decline"
    import_id = parts[2]

    # Check if import exists
    if import_id not in pending_imports:
        await query.edit_message_text(
            "❌ Import expired or already processed.\n"
            "Please upload the file again.",
            parse_mode="Markdown"
        )
        return

    pending = pending_imports.pop(import_id)
    transactions = pending["transactions"]
    sender_name = pending["sender_name"]
    expected_total = pending["expected_total"]
    charge_date = pending.get("charge_date", "")

    if action == "decline":
        await query.edit_message_text(
            "❌ *Import Cancelled*\n\n"
            f"Discarded {len(transactions)} transactions.\n"
            "No data was saved.",
            parse_mode="Markdown"
        )
        return

    # action == "approve"
    saved_count = save_transactions_to_db(transactions, sender_name, charge_date)
    total = sum(tx["amount"] for tx in transactions)
    category_summary, validation_msg, _ = build_import_summary(transactions, expected_total)

    # Log the activity
    log_activity(
        action="file_import",
        source="telegram",
        details=f"Imported by {sender_name} (approved with mismatch)",
        record_count=saved_count,
        file_date=charge_date if charge_date else None,
        total_amount=total
    )

    charge_date_msg = f"\nCharge Date: {charge_date}\n" if charge_date else ""
    reply = (
        f"✅ *Bulk Import Approved & Saved*\n\n"
        f"Transactions: {saved_count}\n"
        f"Total: ₪{total:,.2f}{charge_date_msg}\n\n"
        f"*By Category:*\n{category_summary}"
        f"{validation_msg}"
    )
    await query.edit_message_text(reply, parse_mode="Markdown")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle incoming text messages as expense entries."""
    if not update.message or not update.message.text:
        return

    user = update.effective_user
    chat_id = update.effective_chat.id

    # Security check
    if not is_authorized(user.id, chat_id):
        logger.warning(f"Unauthorized message from user ID: {user.id}")
        return

    raw_text = update.message.text
    sender_name = user.first_name or "Family Member"

    # Parse the expense text
    amount, description, cat_id, cat_name = parse_expense_text(raw_text)

    if amount is None:
        await update.message.reply_text(
            "No amount found.\n\n"
            "Format examples:\n"
            "• `Groceries 340`\n"
            "• `Coffee 18`\n"
            "• `Gas 220`",
            parse_mode="Markdown"
        )
        return

    # Persist to database
    with Session(engine) as session:
        expense = Expense(
            amount=amount,
            description=description,
            category_id=cat_id,
            payer=sender_name,
            is_fixed=False,
            source="manual"
        )
        session.add(expense)
        session.commit()

    # Log the activity
    log_activity(
        action="expense_added",
        source="telegram",
        details=f"{description} ({cat_name}) by {sender_name}",
        record_count=1,
        total_amount=amount
    )

    reply = (
        f"*Expense Logged*\n"
        f"Amount: {amount:,.2f}\n"
        f"Category: {cat_name}\n"
        f"Description: {description}\n"
        f"Logged By: {sender_name}"
    )
    await update.message.reply_text(reply, parse_mode="Markdown")


async def run_bot_async():
    """Run the Telegram bot with proper async handling."""
    if not TELEGRAM_BOT_TOKEN:
        logger.warning("TELEGRAM_BOT_TOKEN not set. Bot worker disabled.")
        return

    application = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()

    # Add handlers
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("bulk", bulk_command))
    application.add_handler(CallbackQueryHandler(handle_import_callback, pattern="^import_"))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("Telegram Bot starting long polling...")

    # Initialize the application
    await application.initialize()
    await application.start()

    # Start polling without blocking
    await application.updater.start_polling(drop_pending_updates=True)

    logger.info("Telegram Bot polling started successfully.")

    # Keep running until stopped
    try:
        while True:
            await asyncio.sleep(1)
    except asyncio.CancelledError:
        logger.info("Bot polling cancelled, shutting down...")
    finally:
        await application.updater.stop()
        await application.stop()
        await application.shutdown()


def run_bot_in_thread():
    """
    Run the Telegram bot in a separate thread with its own event loop.
    This is necessary because FastAPI runs its own async loop.
    """
    if not TELEGRAM_BOT_TOKEN:
        logger.warning("TELEGRAM_BOT_TOKEN not set. Bot worker disabled.")
        return None

    def bot_thread():
        # Create a new event loop for this thread
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        try:
            loop.run_until_complete(run_bot_async())
        except Exception as e:
            logger.error(f"Telegram bot error: {e}")
        finally:
            loop.close()

    thread = threading.Thread(target=bot_thread, daemon=True, name="TelegramBotThread")
    thread.start()
    logger.info("Telegram bot thread started.")
    return thread
